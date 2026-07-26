"""
INFORME EXCEL
=============
Genera el libro Excel completo a partir de la base de datos: hojas de datos,
gráficas nativas (barras comparando plana vs homomórfica) y las matrices de
confusión de cada predicción renderizadas como rejilla de color.

Todo se dibuja a partir de los datos de la BD (principio: el Excel se
regenera entero desde SQLite), sin depender de imágenes externas.
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule


AZUL_CABECERA = '4472C4'
BORDE_FINO = Border(*[Side(style='thin', color='D0D0D0')] * 4)


def _es_pct(cabecera):
    c = cabecera.lower()
    return c.endswith('_pct') or c in (
        'accuracy', 'balanced_accuracy', 'error_rate', 'sensibilidad',
        'especificidad', 'ppv', 'npv', 'roc_auc', 'precision_macro',
        'precision_micro', 'precision_weighted', 'recall_macro', 'recall_micro',
        'recall_weighted', 'f1_macro', 'f1_micro', 'f1_weighted')


def _formato(cabecera):
    c = cabecera.lower()
    if _es_pct(cabecera):
        return '0.00'
    if c.endswith('_gb') or c in ('mcc', 'cohen_kappa', 'log_loss', 'ratio_balance',
                                  'eficiencia_cpu'):
        return '0.000'
    if c.endswith('_s') or c.endswith('_ms') or c == 'throughput' or 'factor' in c:
        return '0.0000'
    return None


def _escribir_hoja(wb, nombre, cabeceras, filas):
    ws = wb.create_sheet(nombre)
    ws.append(list(cabeceras))
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor=AZUL_CABECERA)
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for fila in filas:
        ws.append([round(v, 4) if isinstance(v, float) else v for v in fila])

    formatos = [_formato(c) for c in cabeceras]
    for j, fmt in enumerate(formatos, start=1):
        if fmt:
            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=j).number_format = fmt

    ws.freeze_panes = 'A2'
    for i, cab in enumerate(cabeceras, start=1):
        anchos = [len(str(cab))] + [len(str(f[i - 1])) for f in filas
                                    if i - 1 < len(f) and f[i - 1] is not None]
        ws.column_dimensions[get_column_letter(i)].width = min(max(anchos or [10]) + 2, 40)
    return ws


def _idx(cabeceras, nombre):
    """Índice 1-based de una columna por su nombre (o None)."""
    return cabeceras.index(nombre) + 1 if nombre in cabeceras else None


def _barras(ws_datos, cabeceras, n_filas, col_cat, cols_series, titulo, eje_y):
    """Crea un gráfico de barras que referencia columnas de una hoja de datos."""
    cats_idx = _idx(cabeceras, col_cat)
    series_idx = [_idx(cabeceras, c) for c in cols_series if _idx(cabeceras, c)]
    if not cats_idx or not series_idx or n_filas < 1:
        return None
    chart = BarChart()
    chart.type = 'col'
    chart.title = titulo
    chart.y_axis.title = eje_y
    chart.x_axis.title = col_cat
    chart.height = 9
    chart.width = 20
    for s_idx in series_idx:
        ref = Reference(ws_datos, min_col=s_idx, max_col=s_idx, min_row=1, max_row=n_filas + 1)
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(Reference(ws_datos, min_col=cats_idx, min_row=2, max_row=n_filas + 1))
    return chart


# ---------------------------------------------------------------------------
# SQL de las hojas de datos
# ---------------------------------------------------------------------------

SQL_EXPERIMENTOS = """
    SELECT id, fecha, dataset_id, dataset_nombre, directorio, arquitectura,
           input_size, hidden_layers, num_clases, pca_features, learning_rate,
           regularizacion, num_semillas, mejor_semilla,
           mejor_val_accuracy AS mejor_val_accuracy_pct,
           n_total, n_train, n_val, n_test, n_features_original,
           distribucion_clases, ratio_balance, notas
    FROM experimentos ORDER BY id"""

SQL_SEMILLAS = """
    SELECT s.id, s.experimento_id, e.dataset_nombre, e.directorio,
           s.semilla, s.val_accuracy AS val_accuracy_pct, s.val_loss
    FROM entrenamientos_semilla s
    LEFT JOIN experimentos e ON e.id = s.experimento_id
    ORDER BY s.experimento_id, s.id"""

SQL_PREDICCIONES = """
    SELECT p.id, p.experimento_id, e.dataset_nombre, e.directorio, e.arquitectura,
           p.fecha, p.tipo, p.n_muestras, p.num_clases,
           p.accuracy AS accuracy_pct, p.balanced_accuracy AS balanced_accuracy_pct,
           p.error_rate AS error_rate_pct,
           p.f1_macro AS f1_macro_pct, p.f1_micro AS f1_micro_pct,
           p.f1_weighted AS f1_weighted_pct,
           p.precision_macro AS precision_macro_pct, p.recall_macro AS recall_macro_pct,
           p.sensibilidad AS sensibilidad_pct, p.especificidad AS especificidad_pct,
           p.ppv AS ppv_pct, p.npv AS npv_pct, p.roc_auc AS roc_auc_pct,
           p.mcc, p.cohen_kappa, p.log_loss,
           p.n_correctas, p.n_incorrectas, p.tp, p.tn, p.fp, p.fn,
           p.tiempo_inferencia AS tiempo_inferencia_s, p.tiempo_por_muestra_ms,
           p.throughput AS throughput_muestras_s,
           p.poly_modulus_degree, p.coeff_mod_bit_sizes, p.global_scale_bits, p.grado_taylor,
           p.rango_activacion, p.poly_coeffs,
           p.batch_size, p.batch_size_efectivo, p.num_workers, p.num_workers_efectivos,
           p.tiempo_contexto AS tiempo_contexto_s, p.tiempo_cifrado AS tiempo_cifrado_s,
           p.tiempo_prediccion AS tiempo_prediccion_s, p.tiempo_descifrado AS tiempo_descifrado_s,
           p.cpu_count_logico, p.cpu_count_fisico, p.torch_threads, p.hilos_pico,
           p.cpu_time_usuario_s, p.cpu_time_sistema_s, p.cpu_time_total_s, p.eficiencia_cpu,
           p.ram_total_wsl_gb, p.ram_limite_gb, p.ram_pico_gb, p.ram_pico_proceso_gb,
           p.ram_inicial_gb, p.ram_media_gb, p.ram_disponible_min_gb, p.swap_pico_gb,
           p.matriz_confusion, p.notas
    FROM predicciones p
    LEFT JOIN experimentos e ON e.id = p.experimento_id
    ORDER BY p.id"""

SQL_METRICAS_CLASE = """
    SELECT mc.id, mc.prediccion_id, mc.experimento_id, e.directorio, mc.tipo,
           mc.clase, mc.precision_pct, mc.recall_pct, mc.f1_pct,
           mc.especificidad_pct, mc.soporte
    FROM metricas_clase mc
    LEFT JOIN experimentos e ON e.id = mc.experimento_id
    ORDER BY mc.prediccion_id, mc.clase"""

# Comparativa: última predicción plana y homomórfica de cada experimento
SQL_COMPARATIVA = """
    SELECT e.directorio,
           pl.accuracy   AS accuracy_plana_pct,
           ho.accuracy   AS accuracy_homomorfica_pct,
           ho.accuracy - pl.accuracy AS delta_accuracy_pct,
           pl.f1_macro   AS f1_macro_plana_pct,
           ho.f1_macro   AS f1_macro_homomorfica_pct,
           pl.tiempo_inferencia AS tiempo_plana_s,
           ho.tiempo_inferencia AS tiempo_homomorfica_s,
           CASE WHEN pl.tiempo_inferencia > 0
                THEN ho.tiempo_inferencia / pl.tiempo_inferencia END AS factor_ralentizacion,
           pl.throughput AS throughput_plana,
           ho.throughput AS throughput_homomorfica,
           ho.ram_pico_gb AS ram_pico_homomorfica_gb,
           ho.eficiencia_cpu AS eficiencia_cpu_homomorfica,
           ho.hilos_pico AS hilos_pico_homomorfica
    FROM experimentos e
    LEFT JOIN predicciones pl ON pl.id =
        (SELECT id FROM predicciones WHERE experimento_id = e.id AND tipo = 'plana'
         ORDER BY id DESC LIMIT 1)
    LEFT JOIN predicciones ho ON ho.id =
        (SELECT id FROM predicciones WHERE experimento_id = e.id AND tipo = 'homomorfica'
         ORDER BY id DESC LIMIT 1)
    WHERE pl.id IS NOT NULL OR ho.id IS NOT NULL
    ORDER BY e.id"""


def _hoja_graficas(wb, ws_comp, cab_comp, n_comp):
    ws = wb.create_sheet('Graficas')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Gráficas comparativas (plana vs homomórfica)'
    ws['A1'].font = Font(bold=True, size=14)
    if n_comp < 1:
        ws['A3'] = 'Aún no hay predicciones que comparar.'
        return

    graficos = [
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['accuracy_plana_pct', 'accuracy_homomorfica_pct'],
                 'Accuracy: plana vs homomórfica', 'Accuracy (%)'), 'A3'),
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['f1_macro_plana_pct', 'f1_macro_homomorfica_pct'],
                 'F1-macro: plana vs homomórfica', 'F1-macro (%)'), 'M3'),
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['factor_ralentizacion'],
                 'Factor de ralentización (homomórfica / plana)', 'x veces'), 'A22'),
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['ram_pico_homomorfica_gb'],
                 'Pico de RAM en predicción homomórfica', 'GB'), 'M22'),
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['throughput_plana', 'throughput_homomorfica'],
                 'Throughput (muestras/s)', 'muestras/s'), 'A41'),
        (_barras(ws_comp, cab_comp, n_comp, 'directorio',
                 ['eficiencia_cpu_homomorfica'],
                 'Eficiencia de CPU (homomórfica)', 'fracción de núcleos'), 'M41'),
    ]
    for chart, ancla in graficos:
        if chart is not None:
            ws.add_chart(chart, ancla)


def _hoja_matrices(store, wb):
    ws = wb.create_sheet('MatricesConfusion')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Matrices de confusión (una por predicción)'
    ws['A1'].font = Font(bold=True, size=14)

    _, filas = store._tabla("""
        SELECT p.id, e.directorio, p.tipo, p.accuracy, p.num_clases, p.matriz_confusion
        FROM predicciones p LEFT JOIN experimentos e ON e.id = p.experimento_id
        WHERE p.matriz_confusion IS NOT NULL
        ORDER BY p.experimento_id, p.tipo, p.id""")

    fila_actual = 3
    for pid, directorio, tipo, accuracy, num_clases, cm_json in filas:
        try:
            cm = json.loads(cm_json)
        except (TypeError, ValueError):
            continue
        if not cm or not isinstance(cm[0], list):
            continue
        n = len(cm)
        etiquetas = (['Negativo', 'Positivo'] if n == 2
                     else [f'Clase {i}' for i in range(n)])

        # Título
        acc = f"{accuracy:.2f}%" if accuracy is not None else "?"
        cel = ws.cell(row=fila_actual, column=1,
                      value=f"{directorio}  ·  {tipo}  ·  accuracy {acc}  (pred #{pid})")
        cel.font = Font(bold=True, size=12, color=AZUL_CABECERA)
        fila_actual += 1

        # Cabecera "Predicho"
        ws.cell(row=fila_actual, column=2, value='Predicho →').font = Font(italic=True)
        for j, etq in enumerate(etiquetas):
            c = ws.cell(row=fila_actual, column=3 + j, value=etq)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
        fila_actual += 1

        primera_fila_datos = fila_actual
        for i, etq in enumerate(etiquetas):
            rc = ws.cell(row=fila_actual, column=2, value=f'Real {etq}')
            rc.font = Font(bold=True)
            for j in range(n):
                c = ws.cell(row=fila_actual, column=3 + j, value=cm[i][j])
                c.alignment = Alignment(horizontal='center')
                c.border = BORDE_FINO
                if i == j:
                    c.font = Font(bold=True)  # diagonal = aciertos
            fila_actual += 1

        # Escala de color sobre los valores de la matriz (blanco -> azul)
        rango = (f"{get_column_letter(3)}{primera_fila_datos}:"
                 f"{get_column_letter(2 + n)}{fila_actual - 1}")
        ws.conditional_formatting.add(rango, ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            end_type='max', end_color='4472C4'))

        # Etiqueta lateral "Real"
        ws.cell(row=primera_fila_datos, column=1, value='Real ↓').font = Font(italic=True)
        fila_actual += 2  # separación entre matrices

    if fila_actual == 3:
        ws['A3'] = 'Aún no hay matrices de confusión guardadas.'


def construir_excel(store, ruta):
    """Construye y guarda el libro Excel completo. `store` es un ResultadosStore."""
    wb = Workbook()
    wb.remove(wb.active)

    # Hojas de datos
    for nombre, sql in (('Experimentos', SQL_EXPERIMENTOS),
                        ('Semillas', SQL_SEMILLAS),
                        ('Predicciones', SQL_PREDICCIONES),
                        ('MetricasClase', SQL_METRICAS_CLASE)):
        cab, filas = store._tabla(sql)
        _escribir_hoja(wb, nombre, cab, filas)

    # Comparativa (base de las gráficas)
    cab_comp, filas_comp = store._tabla(SQL_COMPARATIVA)
    ws_comp = _escribir_hoja(wb, 'Comparativa', cab_comp, filas_comp)

    # Gráficas y matrices
    _hoja_graficas(wb, ws_comp, list(cab_comp), len(filas_comp))
    _hoja_matrices(store, wb)

    # Orden final de pestañas: resumen visual primero
    orden = ['Graficas', 'Comparativa', 'MatricesConfusion', 'Predicciones',
             'MetricasClase', 'Experimentos', 'Semillas']
    wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 99)

    wb.save(ruta)
    return ruta
